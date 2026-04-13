from datetime import datetime
from io import BytesIO
from urllib.parse import quote
import html

from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.deps import get_current_user
from app.db.session import get_db
from app.models import Address, AddressStatus, AssetSnapshot, AssetType, User
from app.services.address_service import (
    batch_delete_addresses,
    batch_generate_addresses,
    export_private_keys,
    get_or_create_default_wallet,
    import_address_private_pairs,
    update_address_tag,
)
from app.services.audit_service import write_audit

router = APIRouter(prefix='/addresses', tags=['addresses'])


class BatchGenerateRequest(BaseModel):
    wallet_id: str | None = None
    count: int = Field(default=100, ge=1, le=5000)
    start_index: int | None = Field(default=None, ge=1)
    tag_prefix: str | None = Field(default=None, max_length=32)


class ExportPrivateKeysRequest(BaseModel):
    status: str = 'ACTIVE'
    tag: str | None = None
    address_ids: list[str] = Field(default_factory=list)


class UpdateTagRequest(BaseModel):
    tag: str | None = Field(default=None, max_length=64)


class BatchDeleteRequest(BaseModel):
    address_ids: list[str] = Field(default_factory=list, min_length=1)


class ImportAddressPairsRequest(BaseModel):
    private_keys: list[str] = Field(default_factory=list, max_length=10000)
    pairs: list[str] = Field(default_factory=list, max_length=10000)
    wallet_id: str | None = None
    tag_prefix: str | None = Field(default=None, max_length=32)


def _compute_col_widths(headers: list[str], rows: list[list]) -> list[int]:
    widths: list[int] = []
    for idx, header in enumerate(headers):
        max_len = len(str(header))
        for row in rows:
            if idx >= len(row):
                continue
            cell = row[idx]
            cell_len = len(str(cell)) if cell is not None else 0
            if cell_len > max_len:
                max_len = cell_len
        # Keep a readable range while still adapting to real content length.
        widths.append(max(10, min(100, max_len + 2)))
    return widths


def _build_xlsx_bytes(*, sheet_name: str, headers: list[str], rows: list[list]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append(row)

    col_widths = _compute_col_widths(headers, rows)
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _build_excel_html_bytes(*, headers: list[str], rows: list[list]) -> bytes:
    col_widths = _compute_col_widths(headers, rows)
    cols_html = []
    for width in col_widths:
        cols_html.append(f'<col style="width:{width}ch;">')

    def td(v: object, is_header: bool = False) -> str:
        tag = 'th' if is_header else 'td'
        txt = html.escape(str(v if v is not None else ''))
        return f'<{tag} style="border:1px solid #d9d9d9;padding:4px;white-space:nowrap;">{txt}</{tag}>'

    header_html = ''.join(td(h, is_header=True) for h in headers)
    body_rows = []
    for r in rows:
        body_rows.append('<tr>' + ''.join(td(v) for v in r) + '</tr>')

    doc = (
        '<html><head><meta charset="utf-8"></head><body>'
        '<table cellspacing="0" cellpadding="0" style="border-collapse:collapse;">'
        + ''.join(cols_html)
        + f'<tr style="font-weight:bold;background:#f5f7fa;">{header_html}</tr>'
        + ''.join(body_rows)
        + '</table></body></html>'
    )
    return ('\ufeff' + doc).encode('utf-8')


def _build_spreadsheet_payload(*, sheet_name: str, headers: list[str], rows: list[list], base_name: str) -> tuple[str, str, bytes]:
    try:
        content = _build_xlsx_bytes(sheet_name=sheet_name, headers=headers, rows=rows)
        return (
            f'{base_name}.xlsx',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            content,
        )
    except ModuleNotFoundError:
        content = _build_excel_html_bytes(headers=headers, rows=rows)
        return (
            f'{base_name}.xls',
            'application/vnd.ms-excel; charset=utf-8',
            content,
        )


def _download_response(*, file_name: str, media_type: str, content: bytes) -> StreamingResponse:
    safe_name = quote(file_name)
    headers = {'Content-Disposition': f"attachment; filename*=UTF-8''{safe_name}"}
    return StreamingResponse(
        BytesIO(content),
        media_type=media_type,
        headers=headers,
    )


@router.post('/batch-generate')
def batch_generate(
    payload: BatchGenerateRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    wallet = get_or_create_default_wallet(db)
    wallet_id = payload.wallet_id or str(wallet.id)

    try:
        rows = batch_generate_addresses(
            db,
            wallet_id=wallet_id,
            count=payload.count,
            start_index=payload.start_index,
            tag_prefix=payload.tag_prefix,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    write_audit(
        db,
        actor_user_id=current_user.id,
        action='ADDRESS_BATCH_CREATE',
        target_type='WALLET',
        target_id=wallet_id,
        ip=request.client.host if request.client else None,
        user_agent=request.headers.get('user-agent'),
        detail={'count': payload.count, 'start_index': payload.start_index, 'tag_prefix': payload.tag_prefix},
    )
    db.commit()

    sample = [{'address_base58': x.address_base58, 'addr_index': x.addr_index} for x in rows[:5]]
    end_index = None
    if rows:
        end_index = max(x.addr_index for x in rows)

    return {
        'created_count': len(rows),
        'range': {'start_index': rows[0].addr_index if rows else None, 'end_index': end_index},
        'sample': sample,
    }


@router.post('/import')
def import_pairs(
    payload: ImportAddressPairsRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    wallet = get_or_create_default_wallet(db)
    wallet_id = payload.wallet_id or str(wallet.id)
    lines = [x for x in (payload.private_keys + payload.pairs) if isinstance(x, str) and x.strip()]
    if not lines:
        raise HTTPException(status_code=400, detail='private_keys 不能为空')

    try:
        result = import_address_private_pairs(
            db,
            wallet_id=wallet_id,
            pairs=lines,
            tag_prefix=payload.tag_prefix,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    write_audit(
        db,
        actor_user_id=current_user.id,
        action='ADDRESS_IMPORT_PRIVATE_KEYS',
        target_type='WALLET',
        target_id=wallet_id,
        ip=request.client.host if request.client else None,
        user_agent=request.headers.get('user-agent'),
        detail={
            'input_count': len(lines),
            'imported_count': result['imported_count'],
            'skipped_count': result['skipped_count'],
            'error_count': result['error_count'],
        },
    )
    db.commit()
    return result


@router.get('')
def list_addresses(
    page: int = 1,
    page_size: int = 20,
    status: str = 'ACTIVE',
    tag: str | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    del current_user
    try:
        status_enum = AddressStatus(status)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail='invalid status') from exc

    query = select(Address).where(Address.status == status_enum)
    if tag:
        query = query.where(Address.tag.like(f'%{tag}%'))

    all_items = db.scalars(query.order_by(Address.addr_index.asc())).all()
    total = len(all_items)
    items = all_items[(page - 1) * page_size : (page - 1) * page_size + page_size]

    addr_ids = [x.id for x in items]
    snaps = db.scalars(select(AssetSnapshot).where(AssetSnapshot.address_id.in_(addr_ids))).all() if addr_ids else []

    snap_map: dict = {}
    for s in snaps:
        key = str(s.address_id)
        snap_map.setdefault(key, {})[s.asset.value] = str(s.balance_dec)

    result = []
    for row in items:
        snap = snap_map.get(str(row.id), {})
        result.append(
            {
                'id': str(row.id),
                'addr_index': row.addr_index,
                'address_base58': row.address_base58,
                'status': row.status.value,
                'tag': row.tag,
                'trx_balance_dec': snap.get(AssetType.TRX.value, '0'),
                'usdt_balance_dec': snap.get(AssetType.USDT_TRC20.value, '0'),
                'energy_balance': int(row.energy_balance or 0),
                'bandwidth_balance': int(row.bandwidth_balance or 0),
                'last_scanned_at': row.last_scanned_at.isoformat() if row.last_scanned_at else None,
            }
        )

    return {'items': result, 'page': page, 'page_size': page_size, 'total': total}


@router.post('/export-private-keys')
def export_keys(
    payload: ExportPrivateKeysRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        items = export_private_keys(
            db,
            status=payload.status,
            tag=payload.tag,
            address_ids=payload.address_ids or None,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    write_audit(
        db,
        actor_user_id=current_user.id,
        action='ADDRESS_EXPORT_PRIVATE_KEYS',
        target_type='ADDRESS',
        target_id='batch',
        ip=request.client.host if request.client else None,
        user_agent=request.headers.get('user-agent'),
        detail={'count': len(items), 'status': payload.status, 'tag': payload.tag},
    )
    db.commit()

    return {'count': len(items), 'items': items}


@router.get('/export-all-addresses-txt')
def export_all_addresses_txt(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = select(Address).where(Address.status == AddressStatus.ACTIVE).order_by(Address.addr_index.asc())
    all_items = db.scalars(query).all()
    lines = [row.address_base58 for row in all_items if row.address_base58]
    file_name = f"addresses_all_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
    media_type = 'text/plain; charset=utf-8'
    content = ('\ufeff' + '\r\n'.join(lines)).encode('utf-8')

    write_audit(
        db,
        actor_user_id=current_user.id,
        action='ADDRESS_EXPORT_ADDRESSES',
        target_type='ADDRESS',
        target_id='batch',
        ip=request.client.host if request.client else None,
        user_agent=request.headers.get('user-agent'),
        detail={'count': len(lines), 'format': 'txt'},
    )
    db.commit()

    return _download_response(file_name=file_name, media_type=media_type, content=content)


@router.get('/export-all-private-keys-xlsx')
def export_all_private_keys_xlsx(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        items = export_private_keys(
            db,
            status='ACTIVE',
            tag=None,
            address_ids=None,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    headers = ['addr_index', 'address_base58', 'tag', 'private_key_hex']
    rows = [[x['addr_index'], x['address_base58'], x.get('tag') or '', x['private_key_hex']] for x in items]
    base_name = f"private_keys_all_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    file_name, media_type, content = _build_spreadsheet_payload(
        sheet_name='private_keys',
        headers=headers,
        rows=rows,
        base_name=base_name,
    )

    write_audit(
        db,
        actor_user_id=current_user.id,
        action='ADDRESS_EXPORT_PRIVATE_KEYS',
        target_type='ADDRESS',
        target_id='batch',
        ip=request.client.host if request.client else None,
        user_agent=request.headers.get('user-agent'),
        detail={'count': len(items), 'status': 'ACTIVE', 'tag': None},
    )
    db.commit()

    return _download_response(file_name=file_name, media_type=media_type, content=content)


@router.patch('/{address_id}/tag')
def patch_tag(
    address_id: str,
    payload: UpdateTagRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        row = update_address_tag(db, address_id=address_id, tag=payload.tag)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc

    write_audit(
        db,
        actor_user_id=current_user.id,
        action='ADDRESS_UPDATE_TAG',
        target_type='ADDRESS',
        target_id=str(row.id),
        ip=request.client.host if request.client else None,
        user_agent=request.headers.get('user-agent'),
        detail={'tag': row.tag},
    )
    db.commit()
    return {'id': str(row.id), 'tag': row.tag}


@router.post('/batch-delete')
def batch_delete(
    payload: BatchDeleteRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        changed = batch_delete_addresses(db, address_ids=payload.address_ids)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    write_audit(
        db,
        actor_user_id=current_user.id,
        action='ADDRESS_BATCH_DELETE',
        target_type='ADDRESS',
        target_id='batch',
        ip=request.client.host if request.client else None,
        user_agent=request.headers.get('user-agent'),
        detail={'requested': len(payload.address_ids), 'changed': changed},
    )
    db.commit()
    return {'requested': len(payload.address_ids), 'changed': changed}
